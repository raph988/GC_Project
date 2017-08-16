# -*- coding: utf-8 -*-
"""
Created on Mon Feb 27 16:24:21 2017

@author: Raph
"""

from utils import _format
from utils import getFromConfig
from shutil import copyfile
from time import sleep


from PySide import QtGui

import openpyxl as pyxl
import win32com.client as wc
import os


NUM_CTR = 0
NUM_CTR_C = 1
NUM_CTR_F = 2

NUM_DLV_CH = 0
NUM_DLV_C = 1
NUM_DLV_F = 2


def usine_parser(sheet, sheet_name):
    
    usines = []
    
    sections = find_sections(sheet)
    
    vendeur = {}
    vendeur['nom'] = sheet.cell(row=1, column=1).value.encode('utf8') #_format(sheet.cell(row=1, column=1).value, set_lower = False, clear = True)
    vendeur['fonction'] = ""
    
    for i in range(2, sheet.max_row, 1):
        cell = sheet.cell(row=i, column=1)
        if is_keyword(cell):
            cell_value = _format(sheet.cell(row=i, column=2).value, set_lower = False, clear = True)
            if type(cell_value).__name__ == 'str' and len(cell_value)>0:
                vendeur['fonction'] = cell_value
#                if len(sections) <= 1:
#                    sections.insert(0, i+1)
                break
    if "SIME DARBY UNIMILLS BV" in vendeur['nom']:
        pass
    s_id = 0
    sections.insert(0, i+1)
    for s_id in range(0, len(sections)-1):
        usine = {}
        areas = find_bounding_rects(sheet, (sections[s_id], sections[s_id+1]))
        if len(areas) < 1 :
            continue
        last_adress = None
        for area in areas:
            keywords = test_keyword(sheet.cell(row=area[0][0], column=area[0][1]).value)
            values = get_values(sheet=sheet, column=area[0][1]+1, row_start=area[0][0], row_end=area[1][0])
            for key in keywords:
            
#                if "SIME DARBY UNIMILLS BV" in vendeur['nom']:
#                    if key == "produits":
#                        print area[0][0], area[1][0]
                
                if is_adress(key):
                    if not isinstance(values[0], list):
                        values = [values, ""]
                    last_adress = values
                    
                if key in usine.keys() and not is_ville(key):
#                    pass
                    #####################################################
                    ##                 A DECOMMENTER                   ##
                    #####################################################
                    # cas unique ou plusieurs adresses ( adresse = [adresse, ville] )
#                    if isinstance(usine[key][0], list):
#                        usine[key].append(values)
#                    else:
                        usine[key] = [usine[key]]
                        usine[key].append(values)
                else:
                    if is_ville(key) and last_adress is not None and last_adress[-1] == "":
                        last_adress[-1] = values[0]
                    else:
                        usine[key] = values
            if usine.has_key("adresse usine") and len(usine["adresse usine"]) > 1 and usine["adresse usine"][1] == "":
                if usine.has_key("adresse facturation") and len(usine["adresse facturation"]) > 1 and usine["adresse usine"][1] != "":
                    usine["adresse usine"][1] = usine["adresse facturation"][1]
                    
        usines.append(usine)
        
    vendeur['usines'] = usines
    
#    if 'COUDENE' in vendeur['nom']:
#        for u in vendeur['usines']:
#            print u
                  
    return vendeur
    

def get_values(sheet, column, row_start, row_end):
    values = []
    for i in range(row_start, row_end+1,1):
        cell_value = sheet.cell(row=i, column=column).value
        if cell_value is not None and cell_value is not None:
            values.append(cell_value)
    return values

        
def test_keyword(word):
    keywords = []
    word = _format(word, clear=True, set_lower = True)
    
    if is_facturation_adresse(word):
        keywords.append("adresse facturation")
        
    if is_usine_adresse(word):
        keywords.append("adresse usine")
    
    if is_siege_adresse(word):
        keywords.append("siege")
        
    if is_produits(word):
        keywords.append("produits")
        
    if len(keywords) < 1:
        keywords.append(word)
        
        
    return keywords
                    
def is_adress(key):
    if "adress" in key or "siege" in key:
        return True
    return False
    
def is_facturation_adresse(key):
    key = key.lower()
    if "adress" in key and "factur" in key:
        return True
    return False
    
def is_usine_adresse(key):
    key = key.lower()
    if ("adress" in key and "usine" in key):
        return True
    return False

def is_siege_adresse(key):
    key = key.lower()
    if "siege" in key:
        return True
    return False

def is_ville(key):
    key = key.lower()
    if "ville" in key:
        return True
    return False

def is_produits(key):
    key = key.lower()
    if "produit" in key:
        return True
    return False
    
def find_bounding_rects(sheet, r_limits):
    areas = []
    for line in range(r_limits[0], r_limits[1],1):
        for col in range(1, sheet.max_column+1,1):
#            cell = sheet.cell(row=line, column=1)
            cell = sheet.cell(row=line, column=col)
            if is_keyword(cell):
#                if cell.value == "Produits":
#                    print cell.value
#                areas.append(find_area(sheet, (line, 1), r_limits=r_limits))
                areas.append(find_area(sheet, (line, col), r_limits=r_limits))
    return areas
    
def find_sections(sheet):
    sections = []
    for i in range(1, sheet.max_row+1,1):
        cell = sheet.cell(row=i, column=1)
        if is_section_style(cell):
            sections.append(i)
    sections.append(sheet.max_row+1)
    return sections
    
        
def find_area(sheet, pos, r_limits):
    i, j = pos
    area = [[i,j], [i, j]] # tuple is non mutable type, but list is !
    
    # growth time
    inc_i, inc_j = 1, 1
    while (inc_i != 0 or inc_j != 0):
        if inc_i != 0:
            next_line = list(sheet.cell(row=area[1][0]+1, column=cg) for cg in range(area[0][1], area[1][1]+1, 1)) 
            for cell in next_line:
                if is_keyword(cell) or is_section_style(cell) or area[1][0] == r_limits[1]-1:
#                    print cell.value, is_keyword(cell), is_section_style(cell)
                    inc_i = 0
                    break
                
        if inc_j != 0:
            next_column = list(sheet.cell(row=rg, column=area[1][1]+1) for rg in range(area[0][0], area[1][0]+2, 1))
            for cell in next_column:
                if is_keyword(cell) or is_section_style(cell) or area[1][1] == sheet.max_column-1:
                    inc_j = 0
                    break
        inc_area(area, inc_i, inc_j)
        
#    # wrap time
    inc_i, inc_j = -1, -1
    while (inc_i != 0 or inc_j != 0) and area[1][0] > r_limits[0] and area[1][1] > sheet.min_column:
        
        if inc_i != 0:
            last_line = list(sheet.cell(row=area[1][0], column=cw) for cw in range(area[0][1], area[1][1]+1, 1))
            for cell in last_line:
                if cell.value is not None:
                    inc_i = 0
                    break
                
        if inc_j != 0:
            last_column = list(sheet.cell(row=rw, column=area[1][1]) for rw in range(area[0][0], area[1][0]+1, 1))
            for cell in last_column:
                if cell.value is not None:
                    inc_j = 0
                    break
                    
        inc_area(area, inc_i, inc_j)
    return area

    
def inc_area(area, inc_i, inc_j):
    area[1][0] += inc_i
    area[1][1] += inc_j
#    print "new area is ", area
    
        
        
def is_keyword(cell):
    if cell.font.b == True and cell.font.u == "single" and cell.value is not None:
        if cell.value != "":
            return True
    return False
    
def is_section_style(cell):
    if cell.value is not None and cell.font.color is not None and cell.font.color.theme is not None:
        if str(cell.font.color.theme).isdigit() and cell.font.color.theme != 1:
            if cell.value != "":
                return True
        
    return False



def ReverseExecutionParser():
    
    from classes import ContractsDatabase
    cDB = ContractsDatabase()
    path = getFromConfig("path", "exec")
    
    tmp = makeTMP(path)
    try:
        wb = pyxl.load_workbook(path, read_only=True)
        test = 1
        while test == 1:
            try: 
                wb.save(path)
            except:
                message = 'Fermez le fichier "'+path+'" pour pouvoir poursuivre.'
                res = QtGui.QMessageBox.question(None,'Attention !' , message, QtGui.QMessageBox.Cancel | QtGui.QMessageBox.Yes)
                if res == QtGui.QMessageBox.Cancel:
                    try:
                        wb.close()
                    except:pass
                    return 1
            else:
                wb.close()
                wb = pyxl.load_workbook(path, read_only=True)
                test = 0
            
        
        sheet = wb.get_sheet_by_name("Planning")
        # on va a la derniere ligne puis on de décalle sur la colonne des num de livraison
        
        last_line = 0
        # on va a la derniere ligne
        for last_line in range(1, sheet.max_row+1, 1):
            cell_value = _format(sheet.cell(row=last_line, column=2).value)
            if cell_value == '':
                break
        #• puis on de décalle sur la colonne des num de livraison pour chercher le dernier ajout 
        for last_registered in range(last_line, 0, -1):
            cell_value = _format(sheet.cell(row=last_registered, column=17).value)
            if cell_value != "":
                break
        
        for i in range(last_registered+1, last_line, 1):
            dic = {}
            
            dic["date_appel"] = sheet.cell(row = i, column=2).value
            dic["client_name"] = sheet.cell(row = i, column=3).value
            dic["fournisseur_name"] = sheet.cell(row = i, column=4).value
            cell = sheet.cell(row = i, column=5).value.lower()
            if 'franco' in cell or 'fr' in cell:
                dic["is_franco"] = True
            else:
                dic["is_franco"] = False
            
            dic["ville"] = sheet.cell(row = i, column=6).value
            dic["date_charg_livr"] = sheet.cell(row = i, column=7).value
            dic["horaire_charg_livr"] = sheet.cell(row = i, column=8).value
            dic["quantite"] = sheet.cell(row = i, column=9).value
            dic["marchandise"] = sheet.cell(row = i, column=10).value
            dic["n_ctr"] = sheet.cell(row = i, column=11).value
            dic["ref_fourniss"] = sheet.cell(row = i, column=12).value
            dic["ref_client"] = sheet.cell(row = i, column=13).value
            dic["ref_chargement"] = sheet.cell(row = i, column=14).value
            cell = sheet.cell(row = i, column=15).value or ""
            if "oui" in cell :
                dic["is_confirmed"] = True
            else:
                dic["is_confirmed"] = False
                
            cell = sheet.cell(row = i, column=16).value
            cell = _format(cell, set_lower=True, clear=True)
            if "paye" in cell or "p" in cell:
                dic["is_paid"] = True
            else:
                dic["is_paid"] = False
               
            if dic["n_ctr"] != "":
                num = dic["n_ctr"]
                type_num = NUM_CTR
                ctr_list = cDB.getContractsByNum(num, type_num)
            elif dic["ref_client"] != "" and len(ctr_list) < 1:
                num = dic["ref_client"]
                type_num = NUM_CTR_C
                ctr_list = cDB.getContractsByNum(num, type_num)
            elif dic["ref_fourniss"] != "" and len(ctr_list) < 1:
                num = dic["ref_fourniss"]
                type_num = NUM_CTR_F
                ctr_list = cDB.getContractsByNum(num, type_num)
            if len(ctr_list) < 1:
                message = "Aucun contrat trouvé sous la réf. "+ str(num) + "."
                QtGui.QMessageBox.question(None, 'Attention',  message, QtGui.QMessageBox.Yes)
                continue
            else:
                ctr = ctr_list[0]
                
            dic["n_livr"] = ctr.n_contrat +"-"+ str(len(ctr.livraisons.keys())).zfill(2)
            sheet.cell(row = i, column=17).value = dic["n_livr"]
            ctr.addDelivery(dic)
            if cDB.updateContract(ctr) < 0:
                raise ValueError
                
        wb.save(path)
        wb.close()
    except:
        try:
            wb.close()
            return -1
        except:
            pass
        os.remove(path)
        os.rename(tmp, path)
        return -1
    else:
        os.remove(tmp)
    return 0



# action 0 go to last line
# action > 0 find the line
def GetLineIndex(sheet, action, num = None):
    i = 0
    if action == 0:
        # on va a la derniere ligne
        for i in range(1, sheet.max_row+1, 1):
            cell_value = _format(sheet.cell(row=i, column=2).value)
            if cell_value == '':
                return i
    elif action > 0 and num is not None:
        #on cherche le num de livraison
        for i in range(1, sheet.max_row+1, 1):
            cell_value = _format(sheet.cell(row=i, column=17).value)
            if cell_value == num:
                return i
            

# action 0 new delivery
# action 1 edit delivery
# action 2 edit remove delivery
def ExecutionParser(delivery=None, action = 0):
    
        progressDialog = QtGui.QProgressDialog()
        progressDialog.setAutoClose(False)
        progressDialog.setAutoReset(False)
        label = progressDialog.findChildren(QtGui.QLabel)[0]
        label.setFont(QtGui.QFont("Calibri", 12))
        button = progressDialog.findChildren(QtGui.QPushButton)[0]
        button.hide()
        progressBar = progressDialog.findChildren(QtGui.QProgressBar)[0]
        progressBar.hide()
        
        
        path = getFromConfig("path", "exec")
        tmp = makeTMP(path)
#    try:
        try:
            wb = pyxl.load_workbook(path, read_only=False)
        except:return -1
    
        test = 1
        cpt = 0
        while test == 1 and cpt != 3:
            try: 
                wb.save(path)
            except:
                message = 'Fermez le fichier "'+path+'" pour pouvoir poursuivre.'
                QtGui.QMessageBox.question(None,'Attention !' , message, QtGui.QMessageBox.Yes)
                cpt += 1
                sleep(0.5)
            else:
                cpt += 1
                test = 0
                wb = pyxl.load_workbook(path, read_only=False)
        if cpt == 3:
            try:
                wb.close()
            except: pass
            else:
                raise ValueError
            
        
        progressDialog.setWindowTitle(u"Édition en cours...")
        text = u"\n\nÉdition du document \n"+path
        progressDialog.setLabelText(text)
        progressDialog.show()
        QtGui.QApplication.processEvents()
        
        
        sheet = wb.get_sheet_by_name("Planning")
        
        i = 0
        if action == 0:
            # on va a la derniere ligne
            i = GetLineIndex(sheet, action)
            
#            for i in range(1, sheet.max_row+1, 1):
#                cell_value = _format(sheet.cell(row=i, column=2).value)
#                if cell_value == '':
#    #                print i
#                    break
        elif action == 1 or action == 2:
            #on cherche le num de livraison
            if isinstance(delivery, list):
                i = GetLineIndex(sheet, action, delivery[-1].n_livr)
            else:
                i = GetLineIndex(sheet, action, delivery.n_livr)
            
#            for i in range(1, sheet.max_row+1, 1):
#                cell_value = _format(sheet.cell(row=i, column=17).value)
#                if cell_value == delivery.n_livr:
#    #                print i
#                    break
    
        
        if action != 2 and delivery is not None:
            sheet.cell(row = i, column=2).value = delivery.date_appel
            sheet.cell(row = i, column=3).value = delivery.client_name
            sheet.cell(row = i, column=4).value = delivery.fournisseur_name
            if delivery.is_franco: type_livr = "Franco"
            else: type_livr = "Départ"
            sheet.cell(row = i, column=5).value = type_livr
            sheet.cell(row = i, column=6).value = delivery.ville
            sheet.cell(row = i, column=7).value = delivery.date_charg_livr
            sheet.cell(row = i, column=8).value = delivery.horaire_charg_livr
            sheet.cell(row = i, column=9).value = delivery.quantite
            sheet.cell(row = i, column=10).value = delivery.marchandise
            sheet.cell(row = i, column=11).value = delivery.n_ctr
            sheet.cell(row = i, column=12).value = delivery.ref_fourniss
            sheet.cell(row = i, column=13).value = delivery.ref_client
            sheet.cell(row = i, column=14).value = delivery.ref_chargement
            if delivery.is_confirmed: confirmation = "Oui"
            else: confirmation = "Non"
            sheet.cell(row = i, column=15).value = confirmation
            if delivery.is_paid is True: paiement = "Payé"
            else: paiement = ""
            sheet.cell(row = i, column=16).value = paiement
            sheet.cell(row = i, column=17).value = delivery.n_livr
        elif action == 2 and delivery is not None:
            if isinstance(delivery, list):
                for dlv in delivery:
                    i = GetLineIndex(sheet, action, dlv.n_livr)
                    if i is None: continue
                    for j in range(2, 18, 1):
                        sheet.cell(row = i, column=j).value = ""
            else:        
                for j in range(2, 18, 1):
                    print i, j
                    sheet.cell(row = i, column=j).value = ""
    
        wb.save(path)
    
        progressDialog.close()
        QtGui.QApplication.processEvents()
        
        if action == 0:
            message = 'Souhaitez vous ouvrir le fichier Execution ?'
            res = QtGui.QMessageBox.question(None,'Attention !' , message, QtGui.QMessageBox.No | QtGui.QMessageBox.Yes)
            if res == QtGui.QMessageBox.Yes:
                while True:
                    try:
                        openWorkSheet(path, "Planning", readOnly=False)
                    except:
                        message = "Le fichier Execution est en cours d'utilisation : fermez le avant !"
                        QtGui.QMessageBox.question(None,'Attention !' , message, QtGui.QMessageBox.Cancel | QtGui.QMessageBox.Ok)
                        if res == QtGui.QMessageBox.Cancel:
                            break
                    else:
                        break
        
#    except:
#        try:
#            wb.close()
#        except:pass
#        try:
#            os.remove(path)
#            os.rename(tmp, path)
#        except:pass
#        
#        return -1
#    else:
#        os.remove(tmp)
#        return 0
        return 0
        
                


def openWorkSheet(workbook_path, worksheet_name, readOnly=True):
    if os.path.exists(workbook_path):
        xl = wc.Dispatch("Excel.Application")
        xl.Workbooks.Open(Filename=workbook_path, ReadOnly=readOnly)
        xl.Worksheets(worksheet_name).Activate()
        xl.Visible = True
        del xl
    else:
        print"file doesnt exist."

def makeTMP(filename, remove_it=False):
    base, extansion = filename.split(".")
    tmp = base+"_tmp"+"."+extansion
    
    if remove_it:
        os.remove(filename)
    else:
        cpt = 1
        while os.path.isfile(tmp):
            tmp = base+"_tmp"+str(cpt)+"."+extansion
            cpt += 1
        copyfile(filename, tmp)
        return tmp
    

import sys
if __name__ == '__main__':
#    app = QtGui.QApplication(sys.argv)
    fif = "C:/Users/Raph/Documents/Python Scripts/G&Cie/data/fif.xlsx"
#    makeTMP(fif)
#    fia = "C:/Users/Raph/Documents/Python Scripts/G&Cie/data/fia.xlsx"
#    wb = pyxl.load_workbook(fif)
#    sheet = wb.get_sheet_by_name("UNIMILLS")
#    vendeur = usine_parser(sheet, "UNIMILLS")
#    ExecutionParser(None)
#    ReverseExecutionParser()
#    find_area(sheet, (3,1))
#    wb = load_workbook(fia)
#    sheet = wb.get_sheet_by_name("COUDENE")
#    vendeur = parse_usine(sheet, "CARGILL")
#    print vendeur["usines"][0]["adresse"]
    
#    date = datetime.datetime.now()
#    
#    cm = ClientSheetManager()
#    cm.loadSheet("C:/Users/Raph/Documents/Python Scripts/G&Cie/data/aigremont.xlsx", str(date.year))

    print "END"
#    app.exec_()