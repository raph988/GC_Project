#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Created on Mon Feb 27 11:29:46 2017

@author: Raph
"""


import unicodedata
import openpyxl as pyxl
#import pickle
import cPickle as pickle
from xml.dom import minidom
import configparser, ast
import os.path
import datetime, dateutil
from collections import Mapping
from copy import deepcopy
from geopy.geocoders import Nominatim
from random import random
from re import findall


import matplotlib as mpl
from matplotlib.backends.backend_agg import FigureCanvasAgg
from PySide.QtGui import QImage, QPixmap


###########
from os import listdir, remove
from os.path import isfile, join
import time
#########

import hashlib
_HASHLEN = 20

import chardet

local_last_update = ""
ini_path = "./data/config.ini"



#accent = ['é', 'è', 'ê', 'à', 'ù', 'û', 'ç', 'ô', 'î', 'ï', 'â', 'œ'	, '&', '€', '®']
#html_sym = ['&eacute;', '&egrave', '&ecirc;', '&agrave;', '&ugrave;', '&ccedil;', '&ocirc;', '&icirc;', '&iuml;', '&acirc;', '&amp;', '&euro;','&reg;',  

class TexPixmap:
    def __init__(self):
        #---- set up a mpl figure instance ----
        self.fig = mpl.figure.Figure()
        self.fig.set_canvas(FigureCanvasAgg(self.fig))
        self.renderer = self.fig.canvas.get_renderer()
        
        #---- plot the mathTex expression ----
        self.ax = self.fig.add_axes([0, 0, 1, 1])
        self.ax.axis('off')
        self.ax.patch.set_facecolor('none')
    
    def mathTex_to_QPixmap(self, mathTex, fs):
    
    
        self.fig.patch.set_facecolor('none')
        t = self.ax.text(0, 0, mathTex, ha='left', va='bottom', fontsize=fs)
    
        #---- fit figure size to text artist ----
    
        fwidth, fheight = self.fig.get_size_inches()
        fig_bbox = self.fig.get_window_extent(self.renderer)
    
        text_bbox = t.get_window_extent(self.renderer)
    
        tight_fwidth = text_bbox.width * fwidth / fig_bbox.width
        tight_fheight = text_bbox.height * fheight / fig_bbox.height
    
        self.fig.set_size_inches(tight_fwidth, tight_fheight)
    
        #---- convert mpl figure to QPixmap ----
    
        buf, size = self.fig.canvas.print_to_buffer()
        qimage = QImage.rgbSwapped(QImage(buf, size[0], size[1], QImage.Format_ARGB32))
        qpixmap = QPixmap(qimage)
        self.ax.cla()
        
        return qpixmap
             
#def getInfosFromAdress(address):
#    if isinstance(address, list):
#        address = getInlineArray(address)
#    geolocator = Nominatim()
#    infos_str = geolocator.geocode(address).address



class DataManager:
    """
    ###########################
    ### DATABASE MANAGEMENT ###
    ###########################
    """
    def __init__(self):
        self.backup_dir = getFromConfig("path", "backup_dir")
        self.data_dir = getFromConfig("path", "data_dir")
        self.data_status_dir = getFromConfig("path", "data_status")
        
    def createFile(self, filename, txt):
        with open(join(self.data_dir, filename), 'wb') as f:
            f.write(txt)
        
    def removeFile(self, filename):
        try:
            remove(join(self.data_dir, filename))
        except:
            pass
        
    def safe_saving(self, obj, file_name):
        #make backup before saving
        try:
            self.make_backup()
        except: pass
        
        file_name = self.data_dir+file_name+'.data'
        
        s = pickle.dumps(obj, pickle.HIGHEST_PROTOCOL)
        s += hashlib.sha1(s).digest()
        
        with open(file_name, 'wb') as f:
            f.write(s)
        
        return True
    

    def safe_loading(self, file_name, n_try = 0):
        full_path = self.data_dir+file_name+'.data'
        
        with open(full_path, 'rb') as f:
            pstr = f.read()
        
        data, checksum = pstr[:-_HASHLEN], pstr[-_HASHLEN:]
        try:
            if hashlib.sha1(data).digest() != checksum and n_try < 1:
                return self.safe_loading(file_name, n_try = n_try+1)
            elif hashlib.sha1(data).digest() == checksum:
                return pickle.loads(data)
            else:
                return self.load_backup()
        except:
            return self.load_backup()
        
        return None


    def load_backup(self):
        onlyfiles = [f for f in listdir(self.backup_dir) if isfile(join(self.backup_dir, f)) and f.startswith("Backup")]
        datetimes = list(datetime.datetime.strptime(s.replace("Backup", "").replace('.data', ""), '_%d_%m_%y %H') for s in onlyfiles)
        ordered = sorted(datetimes, reverse=True)
            
        index = 0
        while index < len(ordered):
            file_name = self.backup_dir+list(f for f in onlyfiles if ordered[index].strftime("_%d_%m_%y %H") in f)[0]
            try:
                with open(file_name, 'rb') as f:
                    pstr = f.read()
                data, checksum = pstr[:-_HASHLEN], pstr[-_HASHLEN:]
                if hashlib.sha1(data).digest() != checksum:
                    raise ValueError
                return pickle.loads(data)
            except:
                index += 1
        return None


    def make_backup(self):
        current_data = self.safe_loading("contrats")
        file_name = self.backup_dir+"Backup"+"_"+datetime.datetime.now().strftime("%d_%m_%y %H")+'.data'
        s = pickle.dumps(current_data, pickle.HIGHEST_PROTOCOL)
        s += hashlib.sha1(s).digest()
        
        with open(file_name, 'wb') as f:
            f.write(s)
            
        
    def getStatus(self):
        with open(self.data_status_dir, 'r') as f:
            lines = f.readlines()
            
        try:
            time = datetime.datetime.strptime(lines[0].rstrip(), '%d_%m_%y %H_%M_%S.%f')
        except:
            time = datetime.datetime.strptime(lines[0].rstrip(), '%d_%m_%y %H_%M')
        status = lines[1].rstrip()
        return time, status
    
    def dbLocked(self):
        time, status = self.getStatus()
        if status == "locked":
            return True
        return False
    
    def hasToBeUpdated(self, local_time):
        if local_time is None:
            return True
        
        try:
            local_time = datetime.datetime.strptime(local_time.rstrip(), '%d_%m_%y %H_%M_%S.%f')
        except:
            local_time = datetime.datetime.strptime(local_time.rstrip(), '%d_%m_%y %H_%M')
            
        time, status = self.getStatus()
        print "last db update : ", time, ", last db import : ", local_time
        if local_time < time:
            return True
        return False
#        except:
#            return False
        
        
    def unlockDB(self):
        try:
            with open(self.data_status_dir, 'r') as f:
                data = [line.strip() for line in f]
        
            current_time = datetime.datetime.now().strftime("%d_%m_%y %H_%M_%S.%f")
            
            data[0] = current_time+'\n'
            data[1] = 'opened'
        except:
            time.sleep(0.1)
            return self.unlockDB()
        
        with open(self.data_status_dir, 'w') as f:
            f.writelines( data )
            
    def lockDB(self):
        try:
            with open(self.data_status_dir, 'r') as f:
                data = [line.strip() for line in f]
        
            data[0] = data[0]+'\n'
            data[1] = 'locked'
        except:
            time.sleep(0.1)
            return self.lockDB()
        
        with open(self.data_status_dir, 'w') as f:
            f.writelines( data )
        
        
    def clearDB(self):
        data = self.safe_loading('contrats')
        for n in data.keys():
            if data[n] is None:
                del data[n]
        self.safe_saving(data,'contrats')
        
            
            
      
def save_obj(obj, name, backup=False):
    if backup is True:
        path = getFromConfig("path", "backup_dir")
        name = path+name+"_"+datetime.datetime.now().strftime("%d_%m_%y %H")+'.data'
    else:
        path = getFromConfig("path", "data_dir")
        name = path+name+'.data'
    with open(name, 'wb') as f:
        pickle.dump(obj, f, pickle.HIGHEST_PROTOCOL)

def load_obj(name, backup=False):
    if backup is True:
        path = getFromConfig("path", "backup_dir")
        onlyfiles = [f for f in listdir(path) if isfile(join(path, f)) and f.startswith(name)]
        datetimes = list(datetime.datetime.strptime(s.replace(name, "").replace('.data', ""), '_%d_%m_%y %H') for s in onlyfiles)
        ordered = sorted(datetimes, reverse=True)
        
        index = 0
        while index < len(ordered):
            name = path+list(f for f in onlyfiles if ordered[index].strftime("_%d_%m_%y %H") in f)[0]
            try:
                with open(name, 'rb') as f:
                    return pickle.load(f)
            except:
                index += 1
        
        print "Fail loading ", name
        raise ValueError
                
    else:
        path = getFromConfig("path", "data_dir")
        name = path+name+'.data'
        print "Loading ", name
        try:
            with open(name, 'rb') as f:
                return pickle.load(f)
        except : 
            print "Fail loading ", name
            raise ValueError
        
def compareDB(obj1, obj2):
    s1 = pickle.dumps(obj1, pickle.HIGHEST_PROTOCOL)
    checksum1 = hashlib.sha1(s1).digest()
    
    s2 = pickle.dumps(obj2, pickle.HIGHEST_PROTOCOL)
    checksum2 = hashlib.sha1(s2).digest()
    
    if checksum1 == checksum2:
        return True
    return False
    


def chaos_pickle(obj, file, io_error_chance=0, eof_error_chance=0):
    if random < io_error_chance:
        raise IOError("Chaotic IOError")

    if random < eof_error_chance:
        raise EOFError("Chaotic EOFError")

    return pickle.Pickler(obj, file, pickle.HIGHEST_PROTOCOL)
        
    
            
def getLastUpdateTime():
    path = getFromConfig("path", "update_time")
    settings = configparser.ConfigParser()
    settings._interpolation = configparser.ExtendedInterpolation()
    settings.read(path)
    stime = settings.get("last_update", "time")
    return dateutil.parser.parse(stime)



###############################
### END DATABASE MANAGEMENT ###
###############################

    
def _format(cell_value, set_lower = True, clear = False):
    
    
    if type(cell_value).__name__ == 'str':
        maj_accent = ["À", "Æ", "æ", "Â", "Ä", "Ô", "Û", "Ü", "Ù", "Ö", "Ê", "Ë", "É", "È", "Œ", "œ", "Ï", "Î", "Ç"]
        maj_sans_accent = ["A", "AE", "ae", "A", "A", "O", "U", "U", "U", "O", "E", "E", "E", "E", "OE", "oe", "I", "I", "C"]
        accent = ['é', 'è', 'ê', 'à', 'ù', 'û', 'ç', 'ô', 'î', 'ï', 'â']
        sans_accent = ['e', 'e', 'e', 'a', 'u', 'u', 'c', 'o', 'i', 'i', 'a']
        if clear is True:
            for c, s in zip(accent, sans_accent):
                cell_value = cell_value.replace(c, s)
            for c, s in zip(maj_accent, maj_sans_accent):
                cell_value = cell_value.replace(c, s)
#        cell_value = cell_value.encode('utf-8').strip()
    elif type(cell_value).__name__ == 'unicode':
        if clear is True:
            cell_value = unicodedata.normalize('NFD', cell_value).encode('ascii', 'ignore')
        else:
            cell_value = cell_value.encode('utf-8').strip()
    elif type(cell_value).__name__ == 'NoneType':
        return ''
    else :
        return cell_value
    
    if set_lower:
        cell_value = cell_value.lower()
        
    try:
        cell_value = cell_value.replace('’', "'")
    except:
        pass
    
    return cell_value
    

def convert_encoding(data, new_coding = 'UTF-8'):
    encoding = chardet.detect(data)['encoding']

    if new_coding.upper() != encoding.upper():
        data = data.decode(encoding, data).encode(new_coding)

    return data
        
def compare_values(v1, v2, just_contains = False):
    if type(v1).__name__ == type(v2).__name__:
        v1 = v1.lower()
        v2 = v2.lower()
        if not just_contains:
            if v1 == v2:
                return True
        else:
            if type(v1).__name__ == 'str':
                if v2 in v1:
                    return True
    return False
         
            
def get_keys(sheet, column, row_start):
    index = 2
    keys = []
    cell_value = sheet.cell(row=index, column=column).value
    while type(cell_value).__name__ =='unicode' and index < sheet.max_row+1:
        keys.append([_format(cell_value, clear = True)])
        index = index + 1
        cell_value = sheet.cell(row=index, column=column).value
    return keys
    
    
def is_key(name, keys):
    for k in keys:
        if k[0].lower() == name.lower():
            return True
    return False

def find_keys(sheet, column, rows, keys, begin = None, end = None):
    keys_founded = []
    tmp_id = -1
    i = 0
    for i in range(rows[0], rows[1]+1, 1):
        cell_value = _format(sheet.cell(row=i, column=column).value)
        if type(cell_value).__name__ == 'str':
            if is_key(cell_value, keys):
                if cell_value.lower() == 'usine':
                    if tmp_id == -1:
                        tmp_id = len(keys_founded)
                    else:
                        keys_founded[tmp_id].append(i-1)
                        tmp_id = len(keys_founded)
                if len(keys_founded) > 0 and keys_founded[-1][0] != 'usine':
                    keys_founded[-1].append(i-1)
                keys_founded.append([cell_value, i])
    if tmp_id != -1:
        keys_founded[tmp_id].append(rows[1])
    keys_founded[-1].append(i)
    return keys_founded


def find_usines(sheet, column = 1, rows = None, keys = None):
    usines_limits = []
    usines_list = []
    last_usine = -1
    i = 0
    r_limits = rows or [1, sheet.max_row]

    for i in range(r_limits[0], r_limits[1]+1, 1):
        cell_value = _format(sheet.cell(row=i, column=column).value)
        if type(cell_value).__name__ == 'str' and cell_value.lower() == 'usine':
            if last_usine != -1:
                usines_limits[-1].append(i-1)
                # les limtes de la premiere usine ont ete trouvées -> on la parse
                usines_list.append(parse_usine(sheet, rows = usines_limits[-1], keys = keys))
            usines_limits.append([i])
            last_usine = i
            
    if len(usines_list) < 1:
        return [parse_usine(sheet, keys = keys)]
                
    usines_list.insert(0, parse_usine(sheet, rows = [1, usines_limits[0][0]], keys = keys))
        
    usines_limits[-1].append(i)
    usines_list.append(parse_usine(sheet, rows = usines_limits[-1], keys = keys))
    
    return usines_list
    

def parse_usine(sheet, column = 1, rows = None, keys = None):
    r_limits = rows or [sheet.min_row, sheet.max_row]
    if keys is None: 
        print("You should specify keys to use when parsing file.")
        return None

    keys_founded = []
    tmp_id = -1
    i = 0
    for i in range(r_limits[0], r_limits[1]+1, 1):
        cell_value = _format(sheet.cell(row=i, column=column).value)
        if type(cell_value).__name__ == 'str' and is_key(cell_value, keys):
                if len(keys_founded) > 0:
                    keys_founded[-1].append(i-1)
                keys_founded.append([cell_value, i])
    if tmp_id != -1:
        keys_founded[tmp_id].append(r_limits[1])
    keys_founded[-1].append(i)
    return keys_founded

       

def get_tva(sheet, rows):
    column_count = sheet.max_column
    r_min, r_max = rows
    
    for i in range(r_min, r_max+1, 1):
        for j in range(1, column_count+1, 1):
            cell_value = _format(sheet.cell(row=i, column=j).value, set_lower=False)
            if 'tva' in cell_value.lower():
                return cell_value.split(':')[1]
    return ""
                        
      
def get_products(sheet, column, row_start, row_end):
    pdts = []
    for i in range(row_start, row_end,1):
        cell_value = _format(sheet.cell(row=i, column=column).value, set_lower=False)
        if cell_value is not None and len(cell_value) > 0:
            pdts.append(cell_value)
    return pdts
    
    
    
def codes_produits(L):
    if len(L)==0: # produit vide
        return [[]]
    else:
        K = codes_produits(L[1:]) #appel récursif 
        return [[x]+y for x in L[0] for y in K] #ajouter tous les éléments du premier ensemble au produit cartésien des autres


def remove_item(array, item):
    try:
        array.remove(item)
    except:
        pass

def has_item(array, item):
    try:
        for i in array:
            if i == item:
                return i
    except:
        pass
    return -1

def _mergeDicts(x, y):
    """Given two dicts, merge them into a new dict as a shallow copy."""
    """ PB : not recursive"""
    z = x.copy()
    z.update(y)
    return z


def mergeDicts(d, other):
    """ update nested dict """
    for k, v in other.items():
        d_v = d.get(k)
        if isinstance(v, Mapping) and isinstance(d_v, Mapping):
            mergeDicts(d_v, v)
        else:
            d[k] = deepcopy(v)
            

def getInlineArray(array):
    text = ""
    if isinstance(array, list) and isinstance(array[0], list):
        adr_list = []
        for adr in array:
            adr_list.append(getInlineArray(adr))
        return adr_list
    else:
        for l in array:
            if len(text) == 0: 
                text = l
            else:
                text = text + '\n' + l
        text = text.encode('utf-8').decode('utf-8').strip()
        return text
    
    
def load_configs(file_name, sheet_name):
    wb = pyxl.load_workbook(file_name)
    sheet = wb.get_sheet_by_name(sheet_name)
    return get_keys(sheet, 1, 2)
    
    
def getCitiesName(country_node):
    city_nodes = country_node.getElementsByTagName('city')
    for city in city_nodes:
        name_node = city.getElementsByTagName('name')[0]
        name = _format(name_node.firstChild.nodeValue, set_lower=True)
        yield name
    
def parseXmlCountries(file_name = None):
    if file_name is None:
        file_name = getFromConfig("path", "geo")
        
    xml_data = minidom.parse(file_name)
    cities_dic = {}

    country_nodes = xml_data.getElementsByTagName('country')

    for c in country_nodes:
        nom_node = c.getElementsByTagName('name')[0]
        country_nom = _format(nom_node.firstChild.nodeValue, set_lower=True)
        cities_name = getCitiesName(c)
        for city in cities_name:
            cities_dic[city.lower()] = country_nom

    save_obj(cities_dic, "Cities_dict" )
    return cities_dic
    
def loadCitiesDic(name = "Cities_dict"):
    try:
        dic = load_obj(name)
    except:
        print "fail to load ", name
        dic = parseXmlCountries()
    return dic


iniConfig = {}
def loadIni():
    print "LOADINI"
    settings = configparser.ConfigParser()
    settings._interpolation = configparser.ExtendedInterpolation()
    settings.read(ini_path, encoding='utf_8_sig')

    for section in settings.sections():
        if str(section) != "signature":
            iniConfig[str(section)] = {}
            for name, value in settings.items(section):
                iniConfig[section][name] = ast.literal_eval(value.encode('utf8').rstrip())

    
def getFromConfig(sectionName, itemName = None):
    if len(iniConfig.keys()) < 1:
        loadIni()
        
    try:
        if itemName is None:
            return iniConfig[sectionName.lower()]
        return iniConfig[sectionName.lower()][itemName.lower()]
    except Exception as e:
        print e
        print "Item not found in ini file."


def number_of_days(month, year):
    date = datetime.date(int(year), int(month), 1)
    next_month = date.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    return (next_month - datetime.timedelta(days=next_month.day)).day
    

def format_num(num):
    if isinstance(num, str) or isinstance(num, unicode):
        num = num.replace(',', '.')
        num = findall(r"[-+]?\d*\.*\d+", num)
        if len(num) > 0:
            num = float(num[0])
        else:
            num = 0.0
    return ('%f' % num).rstrip('0').rstrip('.')

    
if __name__ == '__main__':
#    e = getFromConfig("widgets_info", 'formLayout'.lower())
#    e = getFromConfig("signature").encode('utf-8')
#    data = load_obj('contrats')
#    data = [12, 52, 85, 85, [ 782, 852]]
##    data['17118-01'] = None
#    with open('./data/data.txt', 'wb') as f:
#        json.dump(data, codecs.getwriter('utf-8')(f), ensure_ascii=False, indent=4, sort_keys=True)
        
#    with open('./data/data.txt') as data_file:    
#        data = json.load(data_file)

#    print safe_saving(data, 'contrats')
#    clearDB()
#    load_obj("Backup_", backup=True)
    dm = DataManager()
    print dm.getStatus()
#    print getFromConfig("langages")
#    load_obj("Backup", backup=True)
#    print safe_saving({}, "contrats")
#    print safe_loading("contrats")
#    print getStatus()
#    print number_of_days(02, 2017)
 
#    path = getFromConfig("path", "data_status")
#    print getStatus()
#    saveState()
#    print getStatus()
#    loadState()
#    print getStatus()
#    
#    print getFromConfig("monnaie", 'monnaies')
#    chaos_pickle({0:6, 1:5, 2:4, 3:3, 4:2, 5:1, 6:0}, "./usafe_file.data", eof_error_chance=1)
    print "Done"
    