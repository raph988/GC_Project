# -*- coding: utf-8 -*-
"""
Created on Mon Feb 27 11:18:09 2017

@author: Raph
"""   

from utils import codes_produits
from utils import remove_item
from utils import _format
from utils import has_item
from utils import mergeDicts
from utils import loadCitiesDic
from utils import getFromConfig
from utils import getInlineArray
from utils import DataManager

# tools for db management

from parse_excel import usine_parser 
from parse_excel import ExecutionParser

import Levenshtein
import pycountry
from more_itertools import unique_everseen
import re
from PySide import QtGui, QtCore
from os import listdir, remove
from os.path import isfile, join

import datetime
import time
import threading
import openpyxl as pyxl
import inspect

QtCore.QTextCodec.setCodecForCStrings(QtCore.QTextCodec.codecForName("utf-8"))
        
type_fournisseur = 1
type_acheteur = 0
#geo_datas = loadCitiesDic()

NUM_CTR = 0
NUM_CTR_C = 1
NUM_CTR_F = 2


NUM_DLV_CH = 0
NUM_DLV_C = 1
NUM_DLV_F = 2

livraison = {"date" : "", "marchandise" : "", "quantite" : "", "conteneur": "", "conditions": ""}

class Borg:
    __shared_state = {}
    def __init__(self):
         self.__dict__ = self.__shared_state
         
    
# NEW STYLE CLASS BORG
#class Borg2(object):
#    _state = {}
#    def __new__(cls, *p, **k):
#        self = object.__new__(cls)
#        self.__dict__ = cls._state
#        return self
#
#    def __init__(self, foo= ""):
#        print(foo)
#        
        
        
class Country():
    def __init__(self, country_name):
        try:
            self.country = pycountry.countries.get(name=country_name)
        except:
            print("error for : ",country_name)
            self.country = 0
        
    def getCountryCode(self):
        try:
            return int(self.country.numeric) 
        except:
            return 0
    
    def getCountryName(self):
        try:
            return self.country.name  
        except:
            return None
    
    def getCountryCarCode(self):
        try:
            return self.country.alpha_2
        except:
            return None
        
        
class Contact():
    def __init__(self):
        self.nom = None
        self.tel = None
        self.fonction = None
        
        

class Usine():
    def __init__(self, proprietaire, dic):
        self.proprietaire = proprietaire
#        if "cite marine" in proprietaire.nom.lower():
#            pass
        self.adresse = []
        if dic.has_key(getFromConfig("balises_xlsx", "b_adresse_usine")):
            self.adresse = dic[getFromConfig("balises_xlsx", "b_adresse_usine")]
#        elif dic.has_key("siege"):
#            self.adresse = dic["siege"]
            
        if dic.has_key(getFromConfig("balises_xlsx", "b_adresse_fact")):
            f = dic[getFromConfig("balises_xlsx", "b_adresse_fact")]
        else:
            f = self.proprietaire.adresse_fact
        self.facturation = f
        
        self.country_code = ""
        
#        try:
#            self.ville = dic[getFromConfig("balises_xlsx", "b_ville")][0]
#        except:
#            self.ville = ""
            
        try:
            key = list(k for k in dic.keys() if "tva" in k and "num" in k)
            if len(key) > 0:
                key = key[0]
            self.n_tva = dic[key]
        except:
            self.n_tva = ""
            
        if dic.has_key(getFromConfig("balises_xlsx", "b_marchandise")):
            self.produits = list(m for m in dic[getFromConfig("balises_xlsx", "b_marchandise")])
#            self.produits = list(_format(m, set_lower=True, clear=True) for m in dic[getFromConfig("balises_xlsx", "b_marchandise")])
        else:
            self.produits = []
        
#        self.setCountryCode()
        if len(self.getFact()) <= 0 and len(self.getAdr()) > 0:
            self.facturation = self.adresse
        elif len(self.getAdr()) <= 0 and len(self.getFact()) > 0:
            self.adresse = self.facturation
            
    
    def getAdr(self):
        if len(self.adresse) > 0 and len(self.adresse[0]) > 0:
            return getInlineArray(self.adresse[0])
        return ""
    
    def getVille(self):
        if len(self.adresse) > 1 and len(self.adresse[1]) > 0:
            return self.adresse[1]
        return ""
        
    
    def getFact(self):
        if len(self.facturation) > 0 and len(self.facturation[0]) > 0:
            return getInlineArray(self.facturation[0])
        return ""

#    def setCountryCode(self):
#        country_name = None
#        values = geo_datas.values()
#        c = ""
#        for s in self.adresse:
#            for c in s.split():
#                # on retire tous les caractères autre que lettres
#                c = re.sub(r'[^\x61-\x7a]',r'', c.lower())
#                if geo_datas.has_key(c):
#                    country_name = str(geo_datas[c]).title()
#                elif c in values:
#                    country_name = str(c).title()
#                    
#        if country_name is not None:
#            country = Country(country_name)
#            self.country_code = country.getCountryCode()
#        else:
#            self.country_code = 0
#            print("No country found.")
            
    def __str__(self):
        s = ""
        for attrs, value in list(self.__dict__.items()):
            if attrs == "proprietaire":
                s += "---- "+attrs+" : "+str(value.nom)+"\n"
            else:
                s += "---- "+attrs+" : "+str(value)+"\n"
        return s


class Client():
    
    def __init__(self, dic, is_fournisseur = True):
            self.nom = dic[getFromConfig("balises_xlsx", "b_nom")]
            self.fonction = dic[getFromConfig("balises_xlsx", "b_fonction")]
            
            try:
                self.short_name = dic["short_name"]
            except:
                self.short_name = ""
                
            self.type_client = is_fournisseur
            
            tmp_usine0 = dic[getFromConfig("balises_xlsx", "b_usine_liste")][0]
            
            self.adresse_fact = []
            if tmp_usine0.has_key(getFromConfig("balises_xlsx", "b_adresse_fact")):
                f = tmp_usine0[getFromConfig("balises_xlsx", "b_adresse_fact")]
                self.adresse_fact = f
                
            self.siege = []
            if tmp_usine0.has_key("siege"):
                self.siege = tmp_usine0["siege"]
                
#            elif tmp_usine0.has_key(getFromConfig("balises_xlsx", "b_adresse_usine")):
#                self.adresse = tmp_usine0[getFromConfig("balises_xlsx", "b_adresse_usine")]
            
#            if len(self.adresse) > 0 and isinstance(self.adresse[0], list):
#                self.adresse = list(self.adresse[0])
#                del self.adresse[0]

        

            # ATTENTION #
            # Le siege de l'entreprise correspond à l'usine 0
            self.usines = []
            unparsedUsineList = dic[getFromConfig("balises_xlsx", "b_usine_liste")]
#            if len(unparsedUsineList) > 0:
#                self.siege = Usine(self, unparsedUsineList[0])
#                del unparsedUsineList[0]
#            else:
#                self.siege = None
            for u in unparsedUsineList:
                self.usines.append(Usine(self, u))
                
            self.sheet_path = dic["file_name"]
            
            
            
            
    def getAdrSiege(self):
        if len(self.adresse) < 1:
            return None
            
        if isinstance(self.adresse, list) and isinstance(self.adresse[0], list):
            adresse = []
            for ad in self.adresse:
                adresse.append(getInlineArray(ad))
        else:
            adresse = getInlineArray(self.adresse)
        return adresse
    
    def getFactAdr(self):
        facts = []
        for u in self.usines:
            if len(u.facturation) < 1:
                continue
            if len(u.facturation) > 1 and isinstance(u.facturation[1], list):
                for adr in u.facturation:
                    adr_list = list(i[0] for i in facts)
                    if len(facts) == 0 or getInlineArray(adr[0]) not in adr_list:
                        facts.append([getInlineArray(adr[0]), adr[1], u])
            else:
                adr_list = list(i[0] for i in facts)
                if len(facts) == 0 or getInlineArray(u.facturation[0]) not in adr_list:
                    facts.append([getInlineArray(u.facturation[0]), u.facturation[1], u])
        return facts
            
#    [   [ [u'LESIEUR - Direction Mati\xe8res Premi\xe8res', u'29, Quai Aulagnier', u'92665 Asni\xe8res sur Seine Cedex', u'France'], 
#           u'Asni\xe8res'], 
#        [ [u'SAIPOL', u'Boulevard Maritime', u'76530 GRAND-COURONNE', u'France'], 
#           u'Grand-Couronne']
#    ]

    def getUsineAdr(self):
        adrs = []
        for u in self.usines:
            if len(u.adresse) < 1 or len(u.produits) < 1:
                continue
            if len(u.adresse) > 1 and isinstance(u.adresse[1], list):
                for adr in u.adresse:
#                    print "adr ", adr
                    adr_list = list(i[0] for i in adrs)
                    if len(adrs) == 0 or getInlineArray(adr[0]) not in adr_list:
                        adrs.append([getInlineArray(adr[0]), adr[1], u])
            else:
                adr_list = list(i[0] for i in adrs)
#                print "u.adresse ", u.adresse
                if len(adrs) == 0 or getInlineArray(u.adresse[0]) not in adr_list:
                    adrs.append([getInlineArray(u.adresse[0]), u.adresse[1], u])
#        print "getUsineAdr ", adrs
        return adrs    
                
    def getTvaNumber(self):
        return self.usines[0].n_tva
    
            
    def __str__(self):
        s = ""
        for attrs, value in list(self.__dict__.items()):
            s += "---- "+attrs+" : "+str(value)+"\n"
        return s



class Marchandise():
    def __init__(self, code, nom = None):
        self.nom = nom
        self.code = code
        self.fournisseurs = []
        self.acheteurs = []
    
    

class Market(Borg):
    
    def __init__(self):
        Borg.__init__(self)
        
        if not hasattr(self, 'communicator'):
            self.communicator = Communicator()
            
        if not hasattr(self, 'marchandises'):
            self.marchandises = {}

        if not hasattr(self, 'datas'):
            self.datas = None
            
        if not hasattr(self, 'fournisseurs'):
            self.updateFIF()

        if not hasattr(self, 'acheteurs'):
            self.updateFIA()

            
        if not hasattr(self, 'marchandise_list') or not hasattr(self, 'logements') or not hasattr(self, 'paiements'):
            self.updateConfigLanguages()
            
        if not hasattr(self, 'm_repertory'):
            self.repertory = {}
            
        
            
            
    def updateFIF(self):
        self.fournisseurs = []
        for k in self.marchandises.keys():
            self.marchandises[k].fournisseurs = []
        
        fif_path = getFromConfig("path", "fif")
        wb = pyxl.load_workbook(fif_path)
        
        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            client = None
            try:
                client = usine_parser(sheet, sheet_name)
                client['file_name'] = (fif_path,sheet_name)
                client['short_name'] = sheet_name
            except Exception as e:
                print("Pb occured for sheet "+sheet_name+" in fif")
                pass
            else:
                self.add_client(client, is_fournisseur=True)
#        self.communicator.FIF_updated()
        
    def updateFIA(self):
        self.acheteurs = []
        for k in self.marchandises.keys():
            self.marchandises[k].acheteurs = []
        
        fia_path = getFromConfig("path", "fia")
        wb = pyxl.load_workbook(fia_path)
        
        for sheet_name in wb.get_sheet_names():
            sheet = wb.get_sheet_by_name(sheet_name)
            client = None
            try:
                client = usine_parser(sheet, sheet_name)
                client['file_name'] = (fia_path,sheet_name)
                client['short_name'] = sheet_name
            except Exception as e:
                print("Pb occured for sheet "+sheet_name+" in fia")
                pass
            else:
                self.add_client(client, is_fournisseur=False)
#        self.communicator.FIA_updated()
        
    def updateConfigLanguages(self):
        config_path = getFromConfig("path", "config")
        wb = pyxl.load_workbook(config_path)
        self.updateLogements(wb.get_sheet_by_name(getFromConfig("config_sheetnames", "logements")))
        self.updatePaiements(wb.get_sheet_by_name(getFromConfig("config_sheetnames", "paiements")))
        self.updateMarchandises(wb.get_sheet_by_name(getFromConfig("config_sheetnames", "marchandises")))

    def updateLogements(self, sheet):
        logements = {'fr':[], 'en':[], 'es':[]}
        index = 2
        while index < sheet.max_row+1:
            try:
                cell_value = sheet.cell(row=index, column=1).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            logements['fr'].append(cell_value)
            
            try:
                cell_value = sheet.cell(row=index, column=2).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            logements['en'].append(cell_value)
            
            try:
                cell_value = sheet.cell(row=index, column=3).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            logements['es'].append(cell_value)
            index += 1
        self.logements = logements
            
    def updatePaiements(self, sheet):
        paiement = {'fr':[], 'en':[], 'es':[]}
        index = 2
        while index < sheet.max_row+1:
            try:
                cell_value = sheet.cell(row=index, column=1).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            paiement['fr'].append(cell_value)
            
            try:
                cell_value = sheet.cell(row=index, column=2).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            paiement['en'].append(cell_value)
            
            try:
                cell_value = sheet.cell(row=index, column=3).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            paiement['es'].append(cell_value)
            index += 1
        self.paiements = paiement
            
        
    def updateMarchandises(self, sheet):
        marchandises = {'fr':[], 'en':[], 'es':[]}
        index = 2
        while index < sheet.max_row+1:
            try:
                cell_value = sheet.cell(row=index, column=1).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            marchandises['fr'].append(cell_value)
            
            try:
                cell_value = sheet.cell(row=index, column=2).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            marchandises['en'].append(cell_value)
            
            try:
                cell_value = sheet.cell(row=index, column=3).value
                if not isinstance(cell_value, unicode):
                    cell_value = ""
            except:
                cell_value = ""
            marchandises['es'].append(cell_value)
            index += 1
        self.marchandises_list = marchandises
            
        
    def set_datas(self, datas, generate_all_codes = False):
        self.datas = datas
        
        if generate_all_codes is True:
            all_codes = []
            for dic in self.datas:
                tmp_code = []
                for k,v in dic.items():
                    tmp_code.append(v)
                all_codes.append(tmp_code)
            
            for code in codes_produits(all_codes):
                m = Marchandise(code)
                self.marchandises.append(m)

                
    def add_client(self, dic, is_fournisseur):
        # Create a new client from dictionnary
        new_client = Client(dic, is_fournisseur)
        # and add to client list
        if is_fournisseur is True:
            self.fournisseurs.append(new_client)
        else:
            self.acheteurs.append(new_client)
                    
        # classify lists by marchandises
        for usine in new_client.usines:
            for marchandise in usine.produits:
                code = self.get_code_from_name(marchandise)
#                code = _format(marchandise, True, True)
                
#                if not self.marchandises.has_key(code):
#                    self.marchandises_list[code] = marchandise
                
                if not self.marchandises.has_key(code):
                    self.marchandises[code] = Marchandise(code, nom=marchandise)
                
                if is_fournisseur is True:
                    self.marchandises[code].fournisseurs.append(usine)
                else:
                    self.marchandises[code].acheteurs.append(usine)


            
    def get_client(self, nom = None, is_fournisseur = True):
            
        liste = None
        if is_fournisseur is True:
            liste = self.fournisseurs
        else:
            liste = self.acheteurs
        
        if nom is not None:
            for c in liste:
#                print _format(nom, set_lower=True, clear = True), " in ", _format(c.nom, set_lower=True, clear = True) #_format(nom, clear = True)
                if _format(nom, set_lower=True, clear = True) in _format(c.nom, set_lower=True, clear = True):
                    return c
        else:
            return liste
        print "No client found named ", nom
        return None    
           
            
    def get_clients_from_marchandise(self, full_name, is_fournisseur = True, enlarge = False):
        name = _format(full_name, set_lower=True, clear=True)
        
        clients = []
        if enlarge is True:
            set1 = set(name.split(' '))
            for k, v in self.marchandises.items():
                set2 = set(_format(v.nom, clear=True).split(' '))
#                print "set1 :", set1
#                print "set2 :", set2
                if set1.issubset(set2) or set1.issuperset(set2):
                    if is_fournisseur:
                        clients += v.fournisseurs 
                    else:
                        clients += v.acheteurs 
        else:
            code = self.get_code_from_name(name)
            if self.marchandises.has_key(code):
                if is_fournisseur:
                    return self.marchandises[code].fournisseurs 
                else:
                    return self.marchandises[code].acheteurs 
            
#        codes = self.getPotentialCodesFromName(name)
#        for code in codes:
#            if self.marchandises.has_key(code):
#                if is_fournisseur:
#                    clients = clients + self.marchandises[code].fournisseurs 
#                else:
#                    clients = clients + self.marchandises[code].acheteurs 
#                    
#        clients = list(unique_everseen(clients))
        return clients

          
        
    def getPotentialCodesFromName(self, name):
        code = self.get_code_from_name(name, set_default = False)
        key_list = list(self.marchandises.keys())
        if len(key_list)<1 or len(code) != len(key_list[0]):
            print("Error")
            return []
        
        potential = []
        for key in key_list:
            potential.append([key])
            count = 0
            for i in range(0, len(code), 1):
                if code[i] != '*':
                    if key[i] == code[i]:
                        count = count + 1
            potential[-1].append(count)
        
        total_c = len(code) - code.count('*')
        potential = [p for p in potential if p[1] == total_c]
        potential.sort(key = lambda x:x[1], reverse = True)
        
            
        return [x[0] for x in potential]
        
    
    def getMarchandiseFullName(self, marchandise, lang = 'fr'):
        # code is marchandise name 
        # without blanks, upper nor special characters
        if self.marchandiseExist(marchandise):
            code = self.get_code_from_name(marchandise)
            liste = enumerate(self.marchandises_list['fr'])
            for i, v in liste :
                if code == self.get_code_from_name(v):
                    break
#            index = (i for i, v in enumerate(self.marchandises_list['fr']) if code == self.get_code_from_name(v)).next()
#            index = self.marchandises_list['fr'].index(self.marchandises[self.get_code_from_name(code)].nom)
            return self.marchandises_list[lang][i]
#        if marchandise_list.has_key(code):
#            return marchandise_list[code]
        return ""
        
    def get_code_from_name(self, name, set_default = True):
        name = name.replace(' de ', '').replace(" d'", '').replace(" l'", '').replace(' la ', '').replace(' du ', '').replace(' le ', '').replace(' ', '')
        name = _format(name, clear = True, set_lower=True).lower()
#        code = []
#        try:
#            if self.datas is not None:
#                for i in range(0, len(self.datas), 1):
#                    dic = self.datas[i]
#                    for k,v in dic.items():
#                        tmp = False
#                        if k in name:
#                            code.append(v)
#                            tmp = True
#                            break
#                    if tmp is False:
#                        if set_default is True:
#                            code.append('a') 
#                        else:
#                            code.append('*') 
#        except Exception as e:
#            print(e)
#            return None
#        return ''.join(map(str,code))
#        print name
        return name
        
    def marchandiseExist(self, name):
        code = self.get_code_from_name(name)
#        print " - marchandiseExist - ", name, code
#        print self.marchandises.keys()
        if self.marchandises.has_key(code):
            return True
        return False
        
    def get_marchandise_by_name(self, name):
        code = self.get_code_from_name(name)
        if self.marchandises.has_key(code):
            return self.marchandises[code]
        return None
        
    def getNearestClient(self, text, is_fournisseur = True):
        if is_fournisseur is True:
            client_names = [x.nom for x in self.fournisseurs]
#            client_list = self.fournisseurs
        else:
            client_names = [x.nom for x in self.acheteurs]
#            client_list = self.acheteurs
        sortedList = sorted(client_names, key = lambda name: Levenshtein.distance(name.replace(' ','').lower()[:len(text)], text.lower().replace(" ","")))
#        sortedList = sorted(client_list, key = lambda client: Levenshtein.distance(client.nom.replace(' ','').lower()[:len(text)], text.lower().replace(" ","")))

        return sortedList
    
        
    def getNearestMarchandise(self, text):
        text = self.get_code_from_name(text)
#        marchandise_list = list(m.nom for m in self.marchandises_list)
        marchandise_list = self.marchandises_list['fr']
        sortedList = []
        if marchandise_list.has_key(text):
            sortedList = sorted(marchandise_list, key = lambda name: Levenshtein.distance(self.get_code_from_name(name)[:len(text)], text))
        return sortedList
    
    def getNearestPaiement(self, text):
        text = self.get_code_from_name(text)
        paiement_list = self.paiements['fr']
        sortedList = sorted(paiement_list, key = lambda name: Levenshtein.distance(self.get_code_from_name(name)[:len(text)], text))
        return sortedList
        
    def getNearestLogement(self, text):
        text = self.get_code_from_name(text)
        logement_list = self.logements['fr']
        sortedList = sorted(logement_list, key = lambda name: Levenshtein.distance(self.get_code_from_name(name)[:len(text)], text))
        return sortedList



class ContractsDatabase(Borg): # Borg is needed for signals emittance
    
    def __init__(self):
        Borg.__init__(self)
        
        if not hasattr(self, 'communicator'):
            self.communicator = Communicator()
            
        if not hasattr(self, 'dataManager'):
            self.dataManager = DataManager()
            
        self.local_update_time = None
        
        
        if not hasattr(self, 'ctr_list'):
            self.ctr_list = []
        
        if not hasattr(self, 'contrats'):
            if self.loadDatabase(lock=False, onlyIfNecessary = False) == -1:
                self.contrats = {}
                threading.Thread(target=self.saveDatabase).start()
                print "Any contracts database found"
                
            
    def checkDB(self):
        print " -  checkDB  - "
        db1 = self.contrats.copy()
        
        if self.dataManager.dbLocked():
            return -1
        
        self.loadDatabase(lock=False)
        
        try:
            self.contrats = mergeDicts(db1, self.contrats)
        except:
            print "Fail to merge dicts"
            return -1
        else:
            return 0
        
        
    def loadDatabase(self, lock=True, onlyIfNecessary = True):
#        curframe = inspect.currentframe()
#        calframe = inspect.getouterframes(curframe, 2)
#        print 'loadDatabase called by :', calframe[1][3]

        if onlyIfNecessary is True and self.dataManager.hasToBeUpdated(self.local_update_time) is False:
            # Si la base est déja à jour, on charge rien
            print "      - DB already up to date"
            return 1
            
        
        if self.dataManager.dbLocked():
#            self.communicator.cDB_db_locked()
            time.sleep(1)
            if self.dataManager.dbLocked():
                time.sleep(1)
                if self.dataManager.dbLocked():
                    msgBox = QtGui.QMessageBox()
                    msgBox.setText("La base de donnée semble être verrouillée... Un autre utilisateur intervient-il dessus ?")
                    msgBox.setStandardButtons(QtGui.QMessageBox.Yes | QtGui.QMessageBox.No)
                    res = msgBox.exec_()    
                    if res == QtGui.QMessageBox.No:
                        self.dataManager.unlockDB()
                    else: 
                        return -1
        
        if lock is True: self.dataManager.lockDB()
        try:
            self.contrats = self.dataManager.safe_loading("contrats")
        except:
            if lock is True: self.dataManager.unlockDB()
            print "loadDatabase return -1"
            return -1
        if lock is True: self.dataManager.unlockDB()
        
        self.local_update_time = datetime.datetime.now().strftime("%d_%m_%y %H_%M_%S.%f")
#        self.ctr_list = self.getEveryContracts()
        del self.ctr_list[:]
        if self.contrats is None: return -1
        for n, c in self.contrats.items():
            self.ctr_list.append(c)
#        if self.contrats is None:
#            self.communicator.cDB_fail_loading()
#            return False
        return 0
            
    def saveDatabase(self, mergeIfNecessary = True):
        
        error_occured = 0
        if mergeIfNecessary is True and self.dataManager.hasToBeUpdated(self.local_update_time) is True:
            error_occured += self.checkDB()
            
        self.dataManager.lockDB()
        if self.dataManager.safe_saving(self.contrats, "contrats") is False:
            print "Error while saving contracts database."
            self.communicator.cDB_fail_saving()
            error_occured -= 1
            
        self.dataManager.unlockDB()
        self.communicator.cDB_Updated()
        if error_occured < 0:
            return -1
        else:
            return 0
        
        
                
    def addContract(self, ctr):
        if ctr.n_contrat in list(self.contrats.keys()):
            print "Contract already exists."
            return 
        
        if self.loadDatabase() == -1: return
        
        self.contrats[ctr.n_contrat] = ctr
        self.dataManager.removeFile(ctr.n_contrat+".tmp")
        threading.Thread(target=self.saveDatabase).start()
        
        
    def removeContract(self, ctr = None, n_ctr = None):
        if self.loadDatabase() is False: 
            return -1
        
        if ctr is None:
            if self.contrats.has_key(n_ctr):
                deliveries = self.contrats[n_ctr].getDeliveries()
                if self.contrats[n_ctr].removeDelivery(deliveries) < 0:
                    print("Problem occured while removing deliveries from ctr.")
                    return -1
                del self.contrats[n_ctr]
            else:
                self.clearBDD(test_datas = False)
        else:
            n_ctr = ctr.n_contrat
            return self.removeContract(n_ctr = n_ctr)
            
        self.saveDatabase() # we are already in a thread
        
        try:
            tmp_n_ctr = n_ctr or ctr.n_contrat
            self.dataManager.removeFile(tmp_n_ctr+".tmp")
        except:
            pass
        
    
    def clearBDD(self, test_datas = True):
        if test_datas:
            if self.loadDatabase() == -1: return
            
        for num, ctr in self.contrats.items():
            if ctr is None:
                del self.contrats[num]
        self.saveDatabase() # we are already in a thread
        
        
    def updateContract(self, ctr, wait = False):
        if self.loadDatabase() == -1: return -1
        
        ctr.updateQte()
        ctr.updateSolde()
        self.contrats[ctr.n_contrat] = ctr
        if wait is False:
            threading.Thread(target=self.saveDatabase).start()
            return 0
        else:
            return self.saveDatabase()
#        print "ctr updated :", ctr.n_contrat
            
    def cancelContract(self, ctr):
        self.dataManager.removeFile(ctr.n_contrat+".tmp")
        
        if ctr is not None and self.contrats.has_key(ctr.n_contrat):
            threading.Thread(target=self.removeContract, args=(ctr,None,)).start()
                
    def newContract(self):
        if self.loadDatabase() == -1: return
        
        new_contract = Contract()
        new_contract.n_contrat = self.getNewContractNumber()

        self.dataManager.createFile(new_contract.n_contrat+".tmp", "")
        return new_contract
    
    def confirmContract(self, ctr):
        if self.updateContract(ctr, wait=True) < 0:
            return -1
        self.dataManager.removeFile(ctr.n_contrat+".tmp")
        return 0
        
    def getEveryContracts(self):
#        ctr_list = []
        
        if self.loadDatabase(lock=False) >= 0:
            del self.ctr_list[:]
            for n, c in self.contrats.items():
                self.ctr_list.append(c)
            return self.ctr_list
        else:
            return self.ctr_list
    
    def isCtrLocked(self, ctr):
        if self.contrats.has_key(ctr.n_contrat) and self.contrats[ctr.n_contrat] is None:
            return True
        return False
    
    def getContracts(self, by_year = None, by_client = None, by_fourniss = None, by_marchandise=None, ctr_list = None):
#        l_ctr = ctr_list or self.getEveryContracts()
        if ctr_list is not None:
            l_ctr = ctr_list
        else :
            l_ctr = self.getEveryContracts()
            
            
        if by_year is not None:
            for ctr in list(l_ctr):
                if ctr is None : 
                    continue
                elif int(ctr.date_contrat.split('/')[2]) != by_year:
                    l_ctr.remove(ctr)
            return self.getContracts(by_year=None, by_client=by_client, by_fourniss=by_fourniss, by_marchandise=by_marchandise, ctr_list=l_ctr)
        elif by_marchandise is not None:
            for ctr in list(l_ctr):
                if ctr is None : 
                    continue
                elif ctr.marchandise != by_marchandise:
                    l_ctr.remove(ctr)
            return self.getContracts(by_year=None, by_client=by_client, by_fourniss=by_fourniss, by_marchandise=None, ctr_list=l_ctr)
        elif by_client is not None:
            for ctr in list(l_ctr):
                if ctr is None : 
                    continue
                elif ctr.getClientName().lower() != by_client.nom.lower() or ctr.getClientName(shortest=True).lower() != by_client.short_name.lower():
                    l_ctr.remove(ctr)
            return self.getContracts(by_year=None, by_client=None, by_fourniss=by_fourniss, ctr_list=l_ctr)
        elif by_fourniss is not None:
            for ctr in list(l_ctr):
                if ctr is None : 
                    continue
                else: 
                    if ctr.getFournissName().lower() != by_fourniss.nom.lower() or ctr.getFournissName(shortest=True).lower() != by_fourniss.short_name.lower():
                        l_ctr.remove(ctr)
            return self.getContracts(by_year=None, by_client=None, by_fourniss=None, ctr_list=l_ctr)
        return l_ctr
    
    def getContractsByPeriod(self, by_year = None, anterioriy = True, l_ctr = None):
        if l_ctr is None:
            l_ctr = self.getEveryContracts()
        
        if by_year is None:
            return l_ctr
        
        for ctr in list(l_ctr):
            if ctr is None : continue
            if anterioriy is True:
                if int(ctr.date_contrat.split('/')[2]) > by_year:
                    l_ctr.remove(ctr)
            else:
                if int(ctr.date_contrat.split('/')[2]) < by_year:
                    l_ctr.remove(ctr)
        return l_ctr
                    
                    
    def getNewContractNumber(self):
        # contract number = year + dayNumber + '-' + contract number the day
        date = datetime.datetime.now()
        year = date.strftime("%y")
        dayNumber = date.strftime("%j")
        base = year+dayNumber
        
        tmpfiles = [f.replace('.tmp', '') for f in listdir(self.dataManager.data_dir) if isfile(join(self.dataManager.data_dir, f)) and f.endswith(".tmp")]
        contract_number = ""
        new_ctrNumber = 1
        while True:
            contract_number = base + '-'+str(new_ctrNumber).zfill(2)
            if contract_number not in list(self.contrats.keys()) and contract_number not in tmpfiles:
                break
            new_ctrNumber += 1
#        howManyCtr = sum(1 for x in list(self.contrats.keys()) if str(contract_number) in x)
        
        return contract_number
    
    def getContractByNum(self, num):
        if self.loadDatabase(lock=False) -1:
            return None
        if self.contrats.has_key(num):
            return self.contrats[num]
        return None
            
    def getContractsByNum(self, num, type_num = NUM_CTR):
        if self.loadDatabase(lock=False) == -1:
            return None
        
        ctr_list = []
        for n, ctr in self.contrats.items():
            if ctr is None: continue
            
            if type_num is None:
                if num in n or num in ctr.n_client or num in ctr.n_fourniss:
                    ctr_list.append(ctr)
            else:
                if n is not None and type_num == NUM_CTR:
                    if str(num) in n:
                        ctr_list.append(ctr)
                elif type_num == NUM_CTR_C:
                    if ctr.n_client is not None and num in ctr.n_client:
                        ctr_list.append(ctr)
                elif ctr.n_fourniss is not None and type_num == NUM_CTR_F:
                    if num in ctr.n_fourniss:
                        ctr_list.append(ctr)
                        
        return ctr_list
    
        
    def getDeliveries(self, by_month = None, by_year = None, is_appel_date=True, by_client = None, by_fourniss = None, by_marchandise = None, delivery_list = None):
        if delivery_list is None:
            if by_year is not None:
                l_ctr = self.getContractsByPeriod(by_year=by_year, anterioriy=True)
            else:
                l_ctr = None
            l_ctr = self.getContracts(by_client=by_client, by_fourniss=by_fourniss, by_marchandise=by_marchandise,ctr_list=l_ctr)
            
            delivery_list = []
            if len(l_ctr) > 0:
                for ctr in l_ctr:
                    if ctr is None: continue
                    delivery_list += ctr.getDeliveries()
        else:
            if by_client is not None:
                for dlv in list(delivery_list):
                    ctr = self.getContractByNum(dlv.n_ctr)
                    if ctr.getClientName().lower() != by_client.lower() or ctr.getClientName(shortest=True).lower() != by_client.lower():
                        delivery_list.remove(dlv)
                        
            if by_fourniss is not None:
                for dlv in list(delivery_list):
                    ctr = self.getContractByNum(dlv.n_ctr)
                    if ctr.getFournissName().lower() != by_fourniss.lower() or ctr.getFournissName(shortest=True).lower() != by_fourniss.lower():
                        delivery_list.remove(dlv)
                        
        if by_year is not None:
            for delivery in list(delivery_list):
#                delivery = Livraison()
                if is_appel_date is True: 
                    year = int(delivery.date_appel.split('/')[-1])
                else: 
                    year = int(delivery.date_charg_livr.split('/')[-1])
                
                if year != int(by_year):
                    delivery_list.remove(delivery)
                    
        if by_month is not None:
            for delivery in list(delivery_list):
#                delivery = Livraison()
                if is_appel_date is True: 
                    month = delivery.date_appel.split('/')[-2]
                else: 
                    month = delivery.date_charg_livr.split('/')[-2]
                if month != by_month:
                    delivery_list.remove(delivery)
        
        return delivery_list



    def getDeliveryByNum(self, num, type_num = NUM_DLV_CH):
        deliveries = self.getDeliveries()
        for delivery in list(deliveries):
            if type_num == NUM_DLV_CH: #ref chargement
                if num not in delivery.ref_chargement:
                    deliveries.remove(delivery)
            elif type_num == NUM_DLV_C: # reference client
                if num not in delivery.ref_client:
                    deliveries.remove(delivery)
            elif type_num == NUM_DLV_F: # reference fourniss
                if num not in delivery.ref_fourniss:
                    deliveries.remove(delivery)
        return deliveries
        
    def __str__(self):
        s = ""
        for attrs, value in list(self.__dict__.items()):
            s += "---- "+attrs+" : "+str(value)+"\n"
        return s

        
        
class Contract():
    
    def __init__(self):
        self.n_contrat = None
        self.n_client = ""
        self.n_fourniss = ""
        self.date_contrat = None
        self.usine_vendeur = None
        self.adr_uFourniss = ""
        self.usine_acheteur = None
        self.adr_uClient = ""
        self.is_franco = True
        self.usine_cible = None
        self.adr_uCible = ""
        self.marchandise = None
        self.quantite = None
        self.qte_total = None
        self.unite = None #0: kg, 1: t
        self.livraisons = {}
        self.periode_livraison = None
        self.descr_livraison = None
        self.prix = None
        self.prix_total = 0.0
        self.monnaie = None #0: euro, 1:dollar, 2:pound
        self.paiement = None
        self.courtage = None
        self.logement = None
#        self.oil_market = Market()
        self.conditions = None
        self.notes = ""
        self.pdf_link = None
        self.reste_livraison = None
        self.reste_paiement = None
        self.historique_paiement = {}
        self.updateQte()
        self.updateSolde()
        
    
    def getDeliveries(self):
        return list(data for key, data in self.livraisons.items())
    
    def getDeliveryByNum(self, num):
        return self.livraisons[num]
    
    def getDeliveriesByDate(self, year, month = None):
        dlv_list = []
        for key in self.livraisons.keys():
            dlv = self.livraisons[key]
            dlv_month, dlv_year = dlv.date_charg_livr.split('/')
            if int(dlv_year) == int(year):
                if month is not None and int(dlv_month) == int(month):
                    dlv_list.append(dlv)
                elif month is None:
                    dlv_list.append(dlv)
        return dlv_list
        
    def newDelivery(self, dic):
        new_delivery = Livraison(self.n_contrat, dic)
        new_delivery.client_name = self.getClientName()
        new_delivery.fournisseur_name = self.getFournissName()
        new_delivery.is_franco = self.is_franco
        new_delivery.n_livr = self.getNewNumber()
        
        res = ExecutionParser(new_delivery, action=0)
        self.livraisons[new_delivery.n_livr] = new_delivery
        return res, new_delivery
            
    
    
    def getNewNumber(self):
        i = 0
        num_list = list(self.livraisons.keys())
        num_list = sorted(num_list, key = lambda n: n.split('-')[2])
        for num in num_list:
            decomp = num.split('-')[2]
            if int(decomp) != i:
                break
            else:
                i += 1
        return self.n_contrat +"-"+ str(i).zfill(2)
        
    def ckeckKey(self, dlv):
        key = dlv.n_livr
        if key in list(self.livraisons.keys()):
            dlv.n_livr = self.getNewNumber()
        return dlv
            
    
    def addDelivery(self, dic):
        new_delivery = Livraison(self.n_contrat, dic)
        new_delivery.client_name = self.getClientName()
        new_delivery.fournisseur_name = self.getFournissName()
        if new_delivery.n_livr is None: 
            new_delivery.n_livr = self.getNewNumber()
        self.livraisons[new_delivery.n_livr] = new_delivery
        
    def updateDelivery(self, livraison):
        if ExecutionParser(livraison, action=1) < 0:
            return -1
        
        self.livraisons[livraison.n_livr] = livraison
        self.updateQte()
        self.updateSolde()
        
        
    def removeDelivery(self, livraison):

        if ExecutionParser(livraison, action=2) < 0:
            return -1
    
        if isinstance(livraison, list):
            for dlv in livraison:
                del self.livraisons[dlv.n_livr]
        else:
            del self.livraisons[livraison.n_livr]
        self.updateQte()
        self.updateSolde()
        return 0
            
    
    def confirmDelivery(self, livraison, b):
        if ExecutionParser(self.livraisons[livraison.n_livr], action=1) < 0:
            return -1
        
        self.livraisons[livraison.n_livr].setConfirmed(b)
        self.updateQte()
    
        
    def validatePaiment(self, livraison, d):
        valueToSet = True if d is not None else False
        
        if d is not None:
            self.livraisons[livraison.n_livr].setPaid(valueToSet)
            if ExecutionParser(self.livraisons[livraison.n_livr], action=1) < 0:
                self.livraisons[livraison.n_livr].setPaid(not valueToSet)
                return -1
            else:
                day, month, year = d.split('/')
                self.historique_paiement[year] = {month.zfill(2): {day.zfill(2): livraison.prix}}
        else:
            self.livraisons[livraison.n_livr].setPaid(valueToSet)
            if ExecutionParser(self.livraisons[livraison.n_livr], action=1) < 0:
                self.livraisons[livraison.n_livr].setPaid(not valueToSet)
                return -1
        self.updateSolde()
        
        
    def updateSolde(self):
        print "updateSolde"
        if self.prix is None or self.qte_total is None :
            return
        
        if len(re.findall(r"[-+]?\d*\.*\d+", self.courtage.replace(',','.'))) > 0:
            courtage = float(re.findall(r"[-+]?\d*\.*\d+", self.courtage.replace(',','.'))[0])
        else:
            courtage = 0.0
            
        prix = re.findall(r"[-+]?\d*\.*\d+", self.prix.replace(',','.'))
        if len(prix) > 0:
            prix = float(prix[0])
            if self.unite == "kg": #( tonne )
#                if 'kilo' in self.prix or 'kg' in self.prix:
                prix *= 1000.0
        else:
            prix = 0.0
            
        
        if '%' in self.courtage:
            self.prix_total = self.qte_total*prix*courtage/100
        else:
            self.prix_total = self.qte_total*courtage
            
        self.reste_paiement = self.prix_total
        
        for k, l in self.livraisons.items():
            l_qte = float(re.findall(r"[-+]?\d*\.*\d+", l.quantite.replace(',','.'))[0])
            if l.is_paid is True:
                if '%' in self.courtage:
                    reviens = float(l_qte*prix*courtage)
                else:
                    reviens = float(l_qte*courtage)
                    
                self.livraisons[k].prix = reviens
                self.reste_paiement -= reviens
                
        print "prix :", prix
        print "courtage :", courtage
        print "reste_paiement :", self.reste_paiement
        
        
    def updateQte(self):
        print "updateQte"
        if self.qte_total is None :
            return
        
        tmp = re.findall(r"[-+]?\d*\.*\d+", str(self.qte_total).replace(',','.'))
        if len(tmp) > 0:
            self.qte_total = float(tmp[0])
#            if self.unite == 'kg':
#                self.qte_total /= 1000
#                self.unite = 't'
        else:
            self.qte_total = 0.0
            
        self.reste_livraison = self.qte_total
            
        # initialize cet qui a été fait comme livraison
        for k, l in self.livraisons.items():
            date_charg_livr = l.date_theorique.split('/')
            if len(date_charg_livr) == 0:
                continue
            elif len(date_charg_livr) == 2:
                month_dlv, year_dlv = date_charg_livr
            else:
                month_dlv, year_dlv = date_charg_livr[1:]
            self.periode_livraison[year_dlv][month_dlv]["done"] = 0.0
            
        for k, l in self.livraisons.items():
            date_charg_livr = l.date_theorique.split('/')
            if len(date_charg_livr) == 0:
                continue
            elif len(date_charg_livr) == 2:
                month_dlv, year_dlv = date_charg_livr
            else:
                month_dlv, year_dlv = date_charg_livr[1:]
            if l.is_confirmed is True:
                qte_dlv = float(re.findall(r"[-+]?\d*\.*\d+", l.quantite.replace(',','.'))[0])
                self.periode_livraison[year_dlv][month_dlv]["done"] += qte_dlv
                self.reste_livraison -= qte_dlv
                
    
    def getVilleCible(self):
        try:
            return self.adr_uCible[1].encode('utf8')
        except:
            return ""
        
    def getAdrCible(self):
        try:
            return self.adr_uCible[0].encode('utf8')
        except:
            return ""
        
    ######### Client / Acheteur
    def getAcheteur(self):
        return self.usine_acheteur.proprietaire
    
        
    def getClientName(self, shortest = False):
        try:
            client = self.usine_acheteur.proprietaire
            if client is not None:
                if shortest is False:
                    return client.nom.encode('utf8')
                return client.short_name.encode('utf8')
        except:
            return ""
    
    def getFactClient(self):
        try:
#            adr = self.getAcheteur().getAdrSiege()
#            if adr is None:
            adr = self.usine_acheteur.getFact().encode('utf8')
            return adr
        except:
            return ""
    
    def getAdr_uClient(self):
        try:
            return self.adr_uClient[0].encode('utf8')
        except:
            return ""
        
    def get_uVilleAcheteur(self):
        try:
            return self.adr_uClient[1].encode('utf8')
        except:
            return ""
    
    
    def getTVA_uClient(self):
        try:
            return getInlineArray(self.usine_acheteur.n_tva).encode('utf8')
        except:
            return ""
    
    ######### Fournisseur/ Vendeur
    def getVendeur(self):
        try:
            return self.usine_vendeur.proprietaire
        except:
            return None
        
    def getFournissName(self, shortest = False):
        try:
            fourniss = self.usine_vendeur.proprietaire
            if fourniss is not None:
                if shortest is False:
                    return fourniss.nom.encode('utf8')
                return fourniss.short_name.encode('utf8')
        except:
            return ""
        
    def getAdr_uFourniss(self):
        try:
            return self.adr_uFourniss[0].encode('utf8')
        except:
            return ""
        
    def get_uVilleVendeur(self):
        try:
            return self.adr_uFourniss[1].encode('utf8')
        except:
            return ""
    
    
    def getFactFourniss(self):
        try:
            return self.usine_vendeur.getFact().encode('utf8')
        except:
            return ""
        
    def getTVA_U_Fourniss(self):
        try:
            return getInlineArray(self.usine_vendeur.n_tva)
        except:
            return ""
        
    def __str__(self):
        s = ""
        for attrs, value in list(self.__dict__.items()):
            s += "---- "+attrs+" : "+str(_format(value))+"\n"
        return s
        

class Livraison:
    def __init__(self, n_ctr, dic):
        
        if dic.has_key("n_livr"):
            n_livr = dic["n_livr"]
        else:
            self.n_livr = None
        if dic.has_key("client_name"):
            self.client_name = dic["client_name"]
        else:
            self.client_name = None
        if dic.has_key("fournisseur_name"):
            self.fournisseur_name = dic["fournisseur_name"]
        else:
            self.fournisseur_name = None
            
        if dic.has_key("is_franco"):
            self.is_franco = dic["is_franco"]
        else:
            self.is_franco = None
        
        self.n_ctr = n_ctr
        
        try:
            self.date_appel = dic["date_appel"].strftime("%d/%m/%Y")
        except:
            self.date_appel = dic["date_appel"]
        
        if "date_theorique" in dic:
            self.date_theorique = dic["date_theorique"]
        else:
            self.date_theorique = self.date_appel
        
        try:
            self.date_charg_livr = dic["date_charg_livr"].strftime("%d/%m/%Y")
        except:
            self.date_charg_livr = str(dic["date_charg_livr"])
        self.horaire_charg_livr = dic["horaire_charg_livr"]
        self.ville = dic["ville"]
        self.marchandise = dic["marchandise"]
        self.quantite = str(dic["quantite"])
        self.ref_fourniss = dic["ref_fourniss"]
        self.ref_client = dic["ref_client"]
        self.ref_chargement = dic["ref_chargement"]
        if dic.has_key("is_confirmed") is True:
            self.is_confirmed = dic["is_confirmed"]
        else:
            self.is_confirmed = False
        if dic.has_key("is_paid") is True:
            self.is_paid = dic["is_paid"]
        else:
            self.is_paid = False
            
        self.prix = 0.0
        

    def setConfirmed(self, b=True):
        self.is_confirmed = b
        
    def setPaid(self, b=True):
        self.is_paid = b
    
    def __str__(self):
        s = ""
        for attrs, value in list(self.__dict__.items()):
#            if not isinstance(value, unicode):
#                value = str(value)
            s += "---- "+attrs+" : "+str(_format(value))+"\n"
        return s
    
    
class Communicator(QtCore.QObject):
    s_cDB_updated = QtCore.Signal()
    s_cDB_fail_loading = QtCore.Signal()
    s_cDB_fail_saving = QtCore.Signal()
    s_cDB_db_locked = QtCore.Signal()
    s_FIA_updated = QtCore.Signal()
    s_FIF_updated = QtCore.Signal()
    
    def __init__(self, parent = None):
        QtCore.QObject.__init__(self)
#        super(Communicator, self).__init__(parent)
    
    def cDB_Updated(self):
        self.s_cDB_updated.emit()
    def cDB_fail_loading(self):
        self.s_cDB_fail_loading.emit()
        self.popupMessage("Échec du chargement de la base de contrats. Réessayez !")
    def cDB_fail_saving(self):
        self.s_cDB_fail_saving.emit()
        self.popupMessage("Échec de la sauvegarde de la base de contrats. Réessayez !")
    def cDB_db_locked(self):
        self.s_cDB_db_locked.emit()
        self.popupMessage("La base de données est verrouillée... Réessayez !")
    def FIA_updated(self): # not used for now
        self.s_FIA_updated.emit()
    def FIF_updated(self): # not used for now
        self.s_FIF_updated.emit()
        
    def popupMessage(self, message, is_question = False):
        msgBox = QtGui.QMessageBox()
        msgBox.setText(message)
        if is_question:
            msgBox.setStandardButtons(QtGui.QMessageBox.Ok | QtGui.QMessageBox.Cancel)
        msgBox.setDefaultButton(QtGui.QMessageBox.Ok)
        return msgBox.exec_()
        

if __name__ == "__main__":
#    m = Market()
    print "start"
#    import sys
#    app = QtGui.QApplication(sys.argv)
    cDB = ContractsDatabase()
#    m = Market()
#    for u in m.get_client("L'herbe folle", is_fournisseur = False).usines:
#        print u.produits
#    [[u'BARILLA FRANCE SAS\nImmeuble Horizons\n30, cours de l\u2019\xeele Seguin\n92100 Boulogne-Billancourt', u'Boulogne-Billancourt', <__main__.Usine instance at 0x0000000009C7AF88>], 
#     [u'Barilla France SAS\nAvenue des Bergeries\nParc industriel de la Plaine de l\u2019Ain\n01150 SAINT VULBAS', u'Saint Vulbas', <__main__.Usine instance at 0x0000000009C82D48>], 
#     [u"HARRY'S FRANCE SAS \nzi Malterie, \n36130 MONTIERCHAUME\nFrance", u'Montierchaume', <__main__.Usine instance at 0x000000000A276548>], 
#     [u'BARILLA FRANCE\nLE PATIS\n85440 TALMONT SAINT HILAIRE\nFrance', u'Talmont Saint Hilaire', <__main__.Usine instance at 0x0000000009C82808>], 
#     [u"BARILLA FRANCE\nLieu-dit Parc De La Valle De L'escaut\n59264 Onnaing\nFrance", u'Onnaing', <__main__.Usine instance at 0x000000000A21F088>]]
#    


#    for c in m.fournisseurs:
        
#        print c.nom
#        if 'olenex' in c.nom.lower() :
#            print c.nom
#            for usine in c.usines:
#                print usine.adresse
#    for u in m.get_client("SIME DARBY UNIMILLS BV").usines:
#        print u.produits
#    Client.adresse
#    print cDB.contrats.keys()
#    cDB.removeContract(None, "17131-01")
#    for ctr in cDB.getEveryContracts():
#        print ctr.n_contrat
#        print ctr.get_uVilleVendeur()
        
#        if hasattr(acheteur, "short_name"):
#            delattr(acheteur, "short_name")
##            setattr(acheteur, 'short_name', m.get_client(acheteur.nom, acheteur.type_client).short_name)
#            cDB.updateContract(ctr)
#            
#    for ctr in cDB.getEveryContracts():
#        acheteur = ctr.getAcheteur()
#        if not hasattr(acheteur, "short_name"):
#            print "DON'T HAVEEE"
#        else:
#            print "HAVEEE"
#            


#    for ctr in cDB.getEveryContracts():
#        if ctr is None:
#            cDB.removeContract(ctr)
#        else:
#            if not hasattr(ctr, "descr_livraison") : 
#                setattr(ctr, "descr_livraison", "BLA")
#            
#                cDB.updateContract(ctr)
#            print ctr.descr_livraison
#    cDB.saveDatabase(mergeIfNecessary=False)
            

#    for ctr in cDB.getEveryContracts():
#        if hasattr(ctr, "descr_livraison") : 
#            print ctr.descr_livraison
#        else:
#            print "HAS NOT ! "
#        acheteur = ctr.getAcheteur()
#        m_acheteur = m.get_client(acheteur.nom, acheteur.type_client)
#        if m_acheteur is not None:
#            for k,v in m_acheteur.__dict__.items():
#                if not hasattr(acheteur, k) : 
#                    setattr(acheteur, k, v)
#                    cDB.updateContract(ctr)
##                    print 'ADDDDED'
#                    
#    for ctr in cDB.getEveryContracts():
#        acheteur = ctr.getAcheteur()
#        if not hasattr(acheteur, "short_name"):
#            print "DON'T HAVEEE"
#        else:
#            print "HAVEEE"
                    
#        for attrs, value in list(ctr.__dict__.items()):
#            try:
#                if value is None:
#                    print attrs, " is None for ", ctr.n_contrat
#                elif len(value) < 1:
#                    print attrs, " is [] for ", ctr.n_contrat
#            except:
#                pass
#        print ctr.livraisons[0]
#    for l in cDB.getDeliveries(by_year=None):
#        cDB.remove
#        print l

#    b = Borg2(ContractsDatabase)
    for c in cDB.getEveryContracts():
#        pass
#        if c is None:
#            print "None"
#        else:
#            print c.n_contract
#            del c
#            cDB.removeContract(c)
#    cDB.saveDatabase() # we are already in a thread
        print "ctr ", c.n_contrat
        for n, l in c.livraisons.items():
#            pass
            print n#, l
#            if not hasattr(l, "ctr"):
#                setattr(l, 'ctr', cDB.getContractByNum(l.n_ctr))
#                c.updateDelivery(l)
#                cDB.updateContract(c)
#            c.removeDelivery(l)
#            cDB.updateContract(c)
#        print c.getFactClient()
#    print m.marchandises_list['fr'].index(m.marchandises[m.get_code_from_name(u"Huile colza raffinée")].nom)
#    print list(m.marchandises.keys())
#    print m.marchandises.get("huile d'arachide", "").fournisseurs
#    name = _format(u'Huile colza raffin\xe9e', clear = True, set_lower=True).lower()
#    
#    name = name.replace(' de ', '').replace(" d'", '').replace(" l'", '').replace(' la ', '').replace(' du ', '').replace(' le ', '').replace(' ', '')
#    print name   
#    print m.getMarchandiseFullName(u'Huile colza raffin\xe9e', 'fr')
#    print m.get_clients_from_marchandise("Huile d'arachide", True, True)
#    print m.logements['en']
#    print m.marchandiseExist("Beurre de karité biologique")


    print "end"